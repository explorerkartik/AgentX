import os
import json

def analyze_data(file_path: str, question: str = "Summarize this data") -> str:
    try:
        if not os.path.exists(file_path):
            return f"Error: File '{file_path}' nahi mili."

        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".csv":
            import csv
            with open(file_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                headers = reader.fieldnames

            if not rows:
                return "CSV file empty hai."

            total_rows = len(rows)
            sample = rows[:5]

            output = f"📊 CSV Analysis:\n"
            output += f"   Total Rows: {total_rows}\n"
            output += f"   Columns: {', '.join(headers)}\n\n"
            output += f"   Sample Data (first 5 rows):\n"
            for i, row in enumerate(sample, 1):
                output += f"   Row {i}: {json.dumps(row)}\n"

            # Basic stats for numeric columns
            output += f"\n   📈 Basic Stats:\n"
            for col in headers:
                values = []
                for row in rows:
                    try:
                        values.append(float(row[col]))
                    except:
                        pass
                if values:
                    output += f"   {col}: min={min(values)}, max={max(values)}, avg={round(sum(values)/len(values), 2)}\n"

            return output

        elif ext in [".xlsx", ".xls"]:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))

                total_rows = len(rows)
                sample = rows[:5]

                output = f"📊 Excel Analysis:\n"
                output += f"   Total Rows: {total_rows}\n"
                output += f"   Columns: {', '.join([str(h) for h in headers])}\n\n"
                output += f"   Sample Data (first 5 rows):\n"
                for i, row in enumerate(sample, 1):
                    output += f"   Row {i}: {json.dumps(row, default=str)}\n"

                return output

            except ImportError:
                return "Excel support ke liye 'openpyxl' install karo: pip install openpyxl"

        else:
            return f"Supported formats: CSV, XLSX. Ye file supported nahi: {ext}"

    except Exception as e:
        return f"Data analysis error: {str(e)}"


DATA_ANALYSIS_TOOL = {
    "type": "function",
    "function": {
        "name": "analyze_data",
        "description": "Analyze CSV or Excel files. Get statistics, summaries, and insights from data files.",
        "parameters": {
            "type": "object",
            "properties": {
                "file_path": {
                    "type": "string",
                    "description": "Full path to CSV or Excel file"
                },
                "question": {
                    "type": "string",
                    "description": "What to analyze or find in the data",
                    "default": "Summarize this data"
                }
            },
            "required": ["file_path"]
        }
    }
}